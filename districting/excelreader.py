import openpyxl
import sys

# params (percent)
try: 
    max_difference = int(sys.argv[1])
except:
    max_difference = 3
    
try: 
    min_landslide = int(sys.argv[2])
except:
    min_landslide = 85

# vars
data = {}
close_races = []
landslides = []

wb = openpyxl.load_workbook('data/elpo04p020/elpo04p020.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


#for column in sheet.iter_cols(min_col = 5, max_col = 5):
#    for cell in column:
#            print(cell.value)

 
# dump data into a dictionary
for row in sheet.iter_rows():
    temp_data = [cell.value for cell in row]
    try:
        data[temp_data[0]]
    except KeyError:
        data[temp_data[0]] = temp_data[1:]

#for k, v in sorted(data.items()):
#    print(k, v)

for k, v in data.items():
    try:
        dem_temp = float(v[10]) # dem % for this county
        rep_temp = float(v[11]) # rep % for this county
        
        # determine which races were close 
        # (less than max_difference)
        if abs(dem_temp - rep_temp) <= max_difference:
            close_races.append((dem_temp, rep_temp))
        
        # determine which races were landslides 
        # (either party percentage greater than min_landslide)
        if dem_temp > min_landslide or rep_temp > min_landslide \
            and -9999 not in [dem_temp, rep_temp]:
            landslides.append((dem_temp, rep_temp))

    except Exception as e:
        pass
        #print(str(e))

        
#for race in close_races:
#    print(race)
    
print("total number of close races: " + str(len(close_races)))
print("total number of landslides: " + str(len(landslides)))
